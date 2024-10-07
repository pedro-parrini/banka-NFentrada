import pandas as pd
import os
from openpyxl import load_workbook

# Função para salvar os dados do DF em um arquivo Excel
def adicionar_df_excel(arquivo_excel, df):
    # Verificar se o arquivo Excel já existe
    if os.path.exists(arquivo_excel):
        # Carregar o workbook existente
        book = load_workbook(arquivo_excel)
                
        # Verificar se a aba 'Produtos' existe
        if 'Registros' in book.sheetnames:
            with pd.ExcelWriter(arquivo_excel, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                # Carregar a aba existente para pegar o número de linhas preenchidas
                sheet = book['Registros']
                startrow = sheet.max_row

                # Adicionar os novos dados após a última linha preenchida
                df.to_excel(writer, sheet_name='Registros', startrow=startrow, startcol=1, index=False, header=startrow == 1)
        else:
            with pd.ExcelWriter(arquivo_excel, engine='openpyxl', mode='a') as writer:
                # Escrever o DataFrame na nova aba 'Produtos', incluindo o cabeçalho
                df.to_excel(writer, sheet_name='Registros', index=False)



##### Modo de Uso da função

#            Montar um dicionário com as informações do DF
#            registro = ({
#                "Loja": loja,
#                "N° da Nota": codigo_nota,
#                "Data de Vencimento": data_vencimento_formatada,
#                "N° do Boleto": codigo_boleto,
#                "Valor Total do boleto": valor_boleto,
#                "Fornecedor": fornecedor,
#            })

#            Converter o dicionário em DF    0bs: a lista é importante para indicar o índice na conversão dicionário -> lista
#            registro = pd.DataFrame([registro])

            # Chamar a função para fazer o registro
#            adicionar_df_excel(excel_file, registro)