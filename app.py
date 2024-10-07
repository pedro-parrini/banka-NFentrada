import streamlit as st
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from email.message import EmailMessage
import smtplib
import shutil


# Configuração inicial da página
st.set_page_config(
    page_title="BANKA",
    page_icon="💰",
    layout="centered",
    initial_sidebar_state="expanded"
)

# Título da Página
st.title("Sistema de Registro de Compras da BANKA")

# Seleção da Loja
loja = st.selectbox(
    "Selecione a loja em que você trabalha:",
    ["Banka BG", "Banka Copacabana", "Banka São Conrado", "Banka Tijuca"]
)

# Definindo o nome do arquivo para salvar os registros
excel_file = "database.xlsx"

# Função para validar o código do boleto
def validar_boleto(boleto):
    # Verifica se o boleto tem 47 ou 48 dígitos
    if len(boleto) not in [47, 48] or not boleto.isdigit():
        return False

    def calcular_dv_bloco(bloco):
        multiplicador = 2
        soma = 0
        for digito in reversed(bloco):
            produto = int(digito) * multiplicador
            soma += produto if produto < 10 else (produto - 9)
            multiplicador = 1 if multiplicador == 2 else 2
        resto = soma % 10
        return 0 if resto == 0 else 10 - resto

    # Extrair blocos do boleto
    blocos = [
        boleto[:9],
        boleto[10:20],
        boleto[21:31],
        boleto[32:47] if len(boleto) == 48 else None,
    ]

    # Extrair dígitos verificadores dos blocos
    dvs = [
        boleto[9],
        boleto[20],
        boleto[31],
        boleto[47] if len(boleto) == 48 else None,
    ]

    # Validar cada bloco
    for i in range(len(blocos) - (1 if blocos[-1] is None else 0)):
        if calcular_dv_bloco(blocos[i]) != int(dvs[i]):
            return False

    return True

# Função para salvar os dados do DF em um arquivo Excel
def adicionar_df_excel(arquivo_excel, df):
    # Verificar se o arquivo Excel já existe
    if os.path.exists(arquivo_excel):
        # Carregar o workbook existente
        book = load_workbook(arquivo_excel)
                
        # Verificar se a aba 'Produtos' existe
        if 'Registros' in book.sheetnames:
            with pd.ExcelWriter(arquivo_excel, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                # Carregar a aba existente para pegar o número de linhas preenchidas
                sheet = book['Registros']
                startrow = sheet.max_row

                # Adicionar os novos dados após a última linha preenchida
                df.to_excel(writer, sheet_name='Registros', startrow=startrow, startcol=1, index=False, header=startrow == 1)
        else:
            with pd.ExcelWriter(arquivo_excel, engine='openpyxl', mode='a') as writer:
                # Escrever o DataFrame na nova aba 'Produtos', incluindo o cabeçalho
                df.to_excel(writer, sheet_name='Registros', index=False)
    else:
        # Criar um novo arquivo Excel com a aba 'Produtos'
        with pd.ExcelWriter(arquivo_excel, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Registros', index=False)

# Função padrão para enviar email
def enviar_email():

    email_destino = ['pedro.parrini@equityrio.com.br', 'financeiro.banka@gmail.com', 'btb.banka@gmail.com']
    email_origem = "pedro.parrini@equityrio.com.br"
    senha_do_email = 'upvz ljbh zszn kipb'

    msg = EmailMessage()
    msg['From'] = email_origem
    msg['Subject']  = f'[Registro de Compra] {loja} - {fornecedor} - {codigo_nota} - {data_vencimento_formatada}'
    msg['To'] = ['pedro.parrini@equityrio.com.br']
#    msg['To'] = email_destino

    mensagem = f''' 

Prezado, Bruno Peniche!
<br><br>
Informo que uma nova compra foi registrada! Seguem as informações para registro e validação:
<br><br>
Unidade: {loja}<br>
Número da NF: {codigo_nota}<br>
Valor do boleto: {valor_boleto}<br>
<b>Número do boleto: {codigo_boleto}</b><br>
Fornecedor: {fornecedor}<br>
<b>Data de Vencimento: {data_vencimento_formatada}</b><br>
<br>
Por favor, prossiga com o agendamento do boleto!
<br><br>
Att.,
<br>
Pedro Vito M. Parrini - Nvestor

'''

    msg.set_content(mensagem, 'html')

    try:

        with open(nota_path, 'rb') as content_file:
            content = content_file.read()
            msg.add_attachment(content, maintype='application', subtype="pdf", 
                            filename=f'NF {codigo_nota}.pdf')
            
        with open(boleto_path, 'rb') as content_file:
            content = content_file.read()
            msg.add_attachment(content, maintype='application', subtype="pdf", 
                            filename=f'Boleto - {codigo_boleto}.pdf')

        with open('database.xlsx', 'rb') as content_file:
            content = content_file.read()
            msg.add_attachment(content, maintype='application', subtype="xlsx", 
                            filename='database.xlsx') 

        with open(xml_path, 'rb') as content_file:
            content = content_file.read()
            msg.add_attachment(content, maintype='application', subtype="xml", 
                            filename=f'XML - {codigo_nota}.xml')     

    except:
        pass

    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
        smtp.login(email_origem, senha_do_email)
        smtp.send_message(msg)

# Criando áreas distintas para cada loja
if loja:
    st.header(f"Área de Registro - {loja}")
    st.write(f"Registre as notas e boletos referentes à {loja}.")

    # Campos para inserção dos dados de notas e boletos
    nota_upload = st.file_uploader("Upload da Nota Fiscal ou Recibo de Compra (PDF ou Foto Escaneada)", type=["pdf"])
    codigo_nota = st.text_input("Número da Nota")
    boleto_upload = st.file_uploader("Upload do Boleto (PDF ou Foto Escaneada)", type=["pdf"])
    codigo_boleto = str(st.number_input("Número do Boleto"))
    xml_upload = st.file_uploader("Upload do XML da Nota Fiscal (opcional)", type=["xml"])
    fornecedor = st.text_input("Fornecedor")
    valor_boleto = st.number_input("Valor Total do Boleto (R$)", min_value=0.0, step=0.01)
    data_vencimento = st.date_input("Data de Vencimento do Boleto")

    # Botão para registrar
    if st.button(f"Registrar Informações - {loja}"):
        # Validações simples antes do registro
        if nota_upload and codigo_nota and boleto_upload and codigo_boleto and fornecedor and valor_boleto and data_vencimento:
            # Validar o código do boleto
            if not validar_boleto(codigo_boleto):
                st.error("Código do boleto inválido. Por favor, verifique os dados.")
            else:
                # Salvar os arquivos inseridos na pasta correta
                nota_path = os.path.join("uploads", nota_upload.name)
                with open(nota_path, "wb") as f:
                    f.write(nota_upload.getbuffer())

                boleto_path = os.path.join("uploads", boleto_upload.name)
                with open(boleto_path, "wb") as f:
                    f.write(boleto_upload.getbuffer())

                xml_path = os.path.join("uploads", xml_upload.name)
                with open(xml_path, "wb") as f:
                    f.write(xml_upload.getbuffer())

                # Organizando os dados para salvar
                data_vencimento_formatada = data_vencimento.strftime("%d/%m/%Y")

                registro = ({
                    "Loja": loja,
                    "N° da Nota": codigo_nota,
                    "Data de Vencimento": data_vencimento_formatada,
                    "N° do Boleto": codigo_boleto,
                    "Valor Total do boleto": valor_boleto,
                    "Fornecedor": fornecedor,
                })

                registro = pd.DataFrame([registro])

                # Salvar os registros no arquivo Excel
                adicionar_df_excel(excel_file, registro)

                # Enviar o email
                enviar_email()

                # Limpar todos os arquivos da pasta uploads
                for filename in os.listdir("uploads"):
                    file_path = os.path.join("uploads", filename)
                    if os.path.isfile(file_path) or os.path.islink(file_path):
                        os.unlink(file_path)  # Remove arquivo ou link
                    elif os.path.isdir(file_path):
                        shutil.rmtree(file_path)  # Remove diretório

                # Informar o usuário que os arquivos foram salvos com sucesso
                st.success("Registro salvo com sucesso!")

        else:
            st.error("Por favor, preencha todos os campos obrigatórios")
